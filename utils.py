import openpyxl
from pathlib import Path
import numpy as np
import matplotlib.pyplot as plt
from typing import Optional, Tuple, Union, List
from datetime import datetime
import os
from openpyxl import Workbook
def excel_to_numpy(
        file_path: Union[str, Path],
        sheet_name: str = None,
        dtype: np.dtype = np.float64
) -> np.ndarray:

    file_path = Path(file_path)

    if not file_path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")
    if file_path.suffix.lower() != '.xlsx':
        raise ValueError(f"文件不是 .xlsx 格式: {file_path}")
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"工作表 '{sheet_name}' 不存在")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        data = []
        first_row = True
        for row in sheet.iter_rows(values_only=True):
            if first_row:
                first_row = False
                continue  # 跳过第一行
            # 跳过第一列并将剩余值转换为列表
            data.append(list(row[1:]))
        # 关闭工作簿
        workbook.close()
        # 转换为 NumPy 数组
        if not data:
            raise ValueError("Excel 文件中没有数据")
        array = np.array(data, dtype=dtype)
        # print(f"成功转换为 {array.shape} 的 NumPy 矩阵")
        return array
    except Exception as e:
        print(f"处理文件时出错: {str(e)}")
        return np.array([])

def grd_to_numpy(file_path):
    from osgeo import gdal
    dataset = gdal.Open(file_path)
    if not dataset:
        raise RuntimeError("无法打开 GRD 文件")

    # 读取第一个波段
    band = dataset.GetRasterBand(1)
    data = band.ReadAsArray()
    return data


def save_grd(array, output_path, x_size=None, y_size=None,
                 x_min=None, y_min=None, pixel_width=None, pixel_height=None):
    """
    将numpy数组保存为GRD文件

    参数:
        array: numpy数组，要保存的栅格数据
        output_path: str, 输出GRD文件的路径
        x_size, y_size: 栅格的列数和行数，如果未提供则使用array的形状
        x_min, y_min: 栅格左上角的坐标，如果未提供则默认为0
        pixel_width, pixel_height: 像元宽度和高度，如果未提供则默认为1

    返回:
        成功保存返回True，否则抛出异常
    """
    from osgeo import gdal, osr

    # 获取数组的尺寸
    if y_size is None and x_size is None:
        y_size, x_size = array.shape
    elif y_size is None:
        y_size = array.shape[0]
    elif x_size is None:
        x_size = array.shape[1]

    # 设置默认地理参考参数
    if x_min is None:
        x_min = 0
    if y_min is None:
        y_min = 0
    if pixel_width is None:
        pixel_width = 1
    if pixel_height is None:
        pixel_height = 1

    # 创建输出驱动
    driver = gdal.GetDriverByName("GTiff")
    # 创建输出数据集
    dataset = driver.Create(
        output_path, x_size, y_size, 1, gdal.GDT_Float32)

    if dataset is None:
        raise RuntimeError(f"无法创建GRD文件: {output_path}")

    geotransform = (x_min, pixel_width, 0, y_min, 0, -pixel_height)
    dataset.SetGeoTransform(geotransform)

    srs = osr.SpatialReference()
    srs.ImportFromEPSG(4326)  # WGS84坐标系
    dataset.SetProjection(srs.ExportToWkt())

    band = dataset.GetRasterBand(1)
    band.WriteArray(array)

    # 刷新缓存，确保数据写入文件
    band.FlushCache()

    # 关闭数据集
    dataset = None
    return True

def save_xlsx(array, output_path):

    wb = Workbook()
    ws = wb.active

    # 获取数组的维度
    rows, columns = array.shape

    # 将数组数据逐行写入 Excel
    for i in range(rows):
        for j in range(columns):
            # 注意：Excel 行和列索引从 1 开始
            ws.cell(row=i+1, column=j+1, value=array[i, j])

    # 保存工作簿
    wb.save(filename=output_path)
    return True

def numpy_to_xlsx(array, output_path, headers=None, sheet_name="Sheet1"):

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 添加表头（如果提供）
    if headers is not None:
        for j, header in enumerate(headers):
            ws.cell(row=1, column=j+1, value=header)

    # 确定数据起始行
    start_row = 2 if headers is not None else 1

    # 获取数组的维度
    rows, columns = array.shape

    # 写入数据
    for i in range(rows):
        for j in range(columns):
            ws.cell(row=i+start_row, column=j+1, value=array[i, j])

    wb.save(filename=output_path)
    return True

def plot_contour(
        matrix: np.ndarray,
        title: str = "data",
        figsize: Tuple[int, int] = (12, 10),
        cmap: str = "viridis",
        levels: Optional[Union[int, List[float]]] = None,
        show_colorbar: bool = False,
        plot_type: str = "filled",
        save_path: Optional[str] = None,
        show_plot: bool = True
) -> None:
    """
    将 NumPy 二维矩阵绘制成等高线图

    参数:
        matrix: 输入的二维 NumPy 矩阵
        title: 图表标题
        figsize: 图表大小
        cmap: 颜色映射名称
        levels: 等高线级别数或自定义级别列表
        show_colorbar: 是否显示颜色条
        plot_type: 图表类型，可选 'filled'（填充等高线）、'contour'（基本等高线）或 '3d'（3D 表面图）
        save_path: 图表保存路径，若为 None 则不保存
        show_plot: 是否显示图表
    """
    # 创建网格坐标
    x = np.arange(matrix.shape[1])
    y = np.arange(matrix.shape[0])
    X, Y = np.meshgrid(x, y)

    # 创建图形
    plt.figure(figsize=figsize)

    if plot_type == 'filled':
        # 填充等高线图
        contour = plt.contourf(X, Y, matrix, levels=levels, cmap=cmap)
        plt.contour(X, Y, matrix, levels=contour.levels, colors='k', linewidths=0.5)
        plt.title(f"{title} (filled)")

    elif plot_type == 'contour':
        # 基本等高线图
        contour = plt.contour(X, Y, matrix, levels=levels, cmap=cmap, linewidths=2)
        plt.clabel(contour, inline=True, fontsize=8)
        plt.title(f"{title} (contour)")

    elif plot_type == '3d':
        # 3D 表面图
        fig = plt.figure(figsize=figsize)
        ax = fig.add_subplot(111, projection='3d')
        surf = ax.plot_surface(X, Y, matrix, cmap=cmap, linewidth=0, antialiased=True)
        ax.set_title(f"{title} (3D)")
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')
        if show_colorbar:
            fig.colorbar(surf, shrink=0.5, aspect=5)
    else:
        raise ValueError(f"不支持的图表类型: {plot_type}，请选择 'filled', 'contour' 或 '3d'")

    # 设置坐标轴标签和颜色条
    if plot_type != '3d':
        plt.xlabel('X')
        plt.ylabel('Y')
        if show_colorbar:
            plt.colorbar(label='nums')

    plt.tight_layout()

    # 保存图表
    if save_path:
        plt.savefig(save_path, dpi=300, bbox_inches='tight')
        # print(f"图表已保存至: {save_path}")

    # 显示图表
    if show_plot:
        plt.show()


def calculate_tick_interval(data_range):
    """根据数据范围计算合适的刻度间隔"""
    if data_range == 0:
        return 1
    # 确定数据范围的数量级
    magnitude = 10 ** np.floor(np.log10(data_range))
    # 计算初步间隔
    preliminary_interval = data_range / 10
    # 确定合适的间隔（1, 2, 5, 10的倍数）
    intervals = [1, 2, 5, 10]
    interval = magnitude
    for i in intervals:
        if i * magnitude >= preliminary_interval:
            interval = i * magnitude
            break
    return interval


def plot_line_chart(data, title="line chart", x_label="x", y_label="y",
                    line_colors=None, markers=None, line_width=2, marker_size=6,
                    show_grid=True, show_labels=False, fig_size=(12, 6), show=True):
    # 确保data是二维列表
    if not isinstance(data[0], list):
        data = [data]
    # 设置默认颜色和标记
    default_colors = ['blue', 'red', 'green', 'purple', 'orange', 'brown', 'pink', 'gray', 'olive', 'cyan']
    default_markers = ['o', 's', '^', 'D', 'v', '*', 'p', 'h', '8', 'x']
    # 处理颜色参数
    if line_colors is None:
        line_colors = [default_colors[i % len(default_colors)] for i in range(len(data))]
    elif not isinstance(line_colors, list):
        line_colors = [line_colors] * len(data)
    # 处理标记参数
    if markers is None:
        markers = [default_markers[i % len(default_markers)] for i in range(len(data))]
    elif not isinstance(markers, list):
        markers = [markers] * len(data)

    plt.figure(figsize=fig_size)

    # 计算所有数据的最小值和最大值
    all_values = [val for sublist in data for val in sublist]
    if not all_values:
        min_value, max_value = 0, 10
    else:
        min_value = min(all_values)
        max_value = max(all_values)

    # 计算数据范围并确定合适的刻度间隔
    data_range = max_value - min_value
    tick_interval = calculate_tick_interval(data_range)

    # 调整y轴范围，使其是刻度间隔的整数倍
    y_min = tick_interval * np.floor(min_value / tick_interval)
    y_max = tick_interval * np.ceil(max_value / tick_interval) + 0.1*(max_value-min_value)

    # 确保y轴范围至少有两个刻度
    if y_max == y_min:
        y_max += tick_interval

    # 绘制每条线
    for i, series in enumerate(data):
        x = np.arange(1, len(series) + 1)
        plt.plot(x, series, f'{markers[i]}-', color=line_colors[i],
                 linewidth=line_width, markersize=marker_size,
                 label=f'iter {i}' if len(data) > 1 else None)

        if show_labels:
            for j, value in enumerate(series):
                plt.annotate(f'{value}', (x[j], value), textcoords='offset points',
                             xytext=(0, 5), ha='center', fontsize=9)

    plt.title(title, fontsize=16)
    plt.xlabel(x_label, fontsize=12)
    plt.ylabel(y_label, fontsize=12)

    # 设置坐标轴范围
    plt.xlim(0, max(len(s) for s in data) + 1 if data else 2)
    plt.ylim(y_min, y_max)

    # 设置刻度
    plt.xticks(np.arange(0, max(len(s) for s in data) + 2 if data else 3, 2))
    plt.yticks(np.arange(y_min, y_max + tick_interval / 2, tick_interval))

    if show_grid:
        plt.grid(True, linestyle='--', alpha=0.7)

    if len(data) > 1:
        plt.legend()

    plt.tight_layout()
    if show:
        plt.show()
    # plt.savefig("custom_chart.png", dpi=300)  # 保存图表

def get_submatrices(matrix, n):
    """
    遍历一个二维 numpy 矩阵，以每个元素为中心提取边长为 n 的子矩阵。
    :param matrix: 输入的二维 numpy 数组
    :param n: 子矩阵的边长，必须为奇数
    :return: 包含所有可提取子矩阵的列表，格式为 ((i, j), submatrix)
    """
    assert n % 2 == 1, "n 必须是奇数"
    pad = n // 2
    # padded_matrix = np.pad(matrix, pad_width=pad, mode='constant', constant_values=0)
    padded_matrix = np.pad(matrix, pad_width=pad, mode='edge')
    submatrices = []
    rows, cols = matrix.shape
    for i in range(rows):
        for j in range(cols):
            submatrix = padded_matrix[i:i+n, j:j+n]
            submatrices.append(((i, j), submatrix))
    return submatrices

def min_mse_average(clip_grids):
    mean_list = []
    msd_list = []
    for grids in clip_grids:
        grids = np.array(grids)
        mean = np.mean(grids)
        msd = np.mean((grids - mean) ** 2)
        mean_list.append(mean)
        msd_list.append(msd)

    min_index = np.argmin(msd_list)
    min_mean = mean_list[min_index]
    min_msd = msd_list[min_index]
    return min_mean, min_msd

def get_current_date_formatted():
    # 获取当前时间
    now = datetime.now()
    # 格式化为 YYYYMMDD 形式
    formatted_date = now.strftime('%Y%m%d-%H%M-%S')
    return formatted_date

def mkdir_if_not_exist(path):
    if not os.path.exists(path):
        os.makedirs(path)